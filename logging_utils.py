# =======================
# File: logging_utils.py
# =======================
"""
Handles logging of processed work orders to CSV and XLSX files.
"""
import csv
import openpyxl
import logging
from datetime import datetime
from pathlib import Path

# Define the headers for the log files
LOG_HEADER = [
    "Broj RNaloga", "Partner", "Aparat", "Serijski broj", "Šifra aparata",
    "Verzija SW", "Šifra pogreške", "Opis pogreške", "Opis posla",
    "Serviser", "Datum", "Potrošni materijal", "Izvorna datoteka"
]

def log_to_csv(data, watched_folder):
    """
    Logs the extracted data to a CSV file in the watched folder.
    The CSV filename is based on the current date.
    """
    try:
        # Get the current date for the filename
        today = datetime.now().strftime('%Y_%m_%d')
        csv_filename = f"{today}.csv"
        csv_filepath = Path(watched_folder) / csv_filename

        # Check if the file exists to determine if we need to write the header
        file_exists = csv_filepath.exists()

        with open(csv_filepath, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')

            if not file_exists:
                writer.writerow(LOG_HEADER)

            # Prepare the row data from the dictionary
            row_data = [
                data.get("work_order_number"),
                data.get("partner"),
                data.get("aparat"),
                data.get("serijski_broj"),
                data.get("sifra_aparata"),
                data.get("verzija_sw"),
                data.get("sifra_pogreske"),
                data.get("opis_pogreske"),
                data.get("opis_obavljenog_posla"),
                data.get("serviser"),
                data.get("datum"),
                data.get("potrosni_materijal"),
                data.get("izvorna_datoteka"),
            ]
            writer.writerow(row_data)

    except Exception as e:
        logging.error(f"Error logging to CSV: {e}")


# --- XLSX Logging ---

CROATIAN_MONTHS = {
    1: "Siječanj", 2: "Veljača", 3: "Ožujak", 4: "Travanj", 5: "Svibanj",
    6: "Lipanj", 7: "Srpanj", 8: "Kolovoz", 9: "Rujan", 10: "Listopad",
    11: "Studeni", 12: "Prosinac"
}

def log_to_excel(data, watched_folder):
    """
    Logs the extracted data to an XLSX file in the watched folder.
    The XLSX filename is based on the current year, and the sheet name on the current month.
    """
    try:
        now = datetime.now()
        year = now.year
        month_name = CROATIAN_MONTHS[now.month]

        xlsx_filename = f"Lista radni nalozi {year}.xlsx"
        xlsx_filepath = Path(watched_folder) / xlsx_filename

        # Load workbook or create a new one
        if xlsx_filepath.exists():
            workbook = openpyxl.load_workbook(xlsx_filepath)
        else:
            workbook = openpyxl.Workbook()
            # Remove the default 'Sheet'
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

        # Get sheet or create a new one
        if month_name in workbook.sheetnames:
            sheet = workbook[month_name]
        else:
            sheet = workbook.create_sheet(title=month_name)
            sheet.append(LOG_HEADER)

        # Prepare the row data
        row_data = [
            data.get("work_order_number"), data.get("partner"), data.get("aparat"),
            data.get("serijski_broj"), data.get("sifra_aparata"), data.get("verzija_sw"),
            data.get("sifra_pogreske"), data.get("opis_pogreske"), data.get("opis_obavljenog_posla"),
            data.get("serviser"), data.get("datum"), data.get("potrosni_materijal"),
            data.get("izvorna_datoteka")
        ]
        sheet.append(row_data)

        workbook.save(xlsx_filepath)

    except Exception as e:
        logging.error(f"Error logging to XLSX: {e}")
